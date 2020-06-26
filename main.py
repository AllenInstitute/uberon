#!/usr/bin/env python
import os
import sys
from openpyxl import load_workbook
from allen_institute_structures import *
from owlready2 import *
from sparql_queries import *
import urllib.request
import json

import networkx as nx
import matplotlib.pyplot as plt

MAPPING_PATH = 'source/one_to_one_mapping.xlsx'
OWL_PATH = 'source/uberon.owl'
MAPPING_SHEET = 'JR_WorkingUberonMapping'
OUTPUT_PATH = 'source/bridge.json'
OUTPUT_GRAPH_PATH = 'source/graph.svg'
MOUSE_OUTPUT_GRAPH_PATH = 'source/mouse.svg'
HUMAN_OUTPUT_GRAPH_PATH = 'source/human.svg'
OWL_URI = "http://purl.obolibrary.org/obo/uberon/releases/2019-11-22/uberon.owl"

MOUSE_API = 'http://api.brain-map.org/api/v2/structure_graph_download/1.json'
BRAINSPAN_API = 'http://api.brain-map.org/api/v2/structure_graph_download/16.json'


def get_report_data():
	print('Loading', MAPPING_PATH)
	mapping_book = load_workbook(MAPPING_PATH, read_only=True, data_only=True)

	# check for existing sheet name
	if MAPPING_SHEET not in mapping_book.sheetnames:
		raise IndexError(MAPPING_PATH + ' does not contain a ' + MAPPING_SHEET + ' sheet')

	return mapping_book[MAPPING_SHEET]

def get_json_from_api(url):
	data = None

	with urllib.request.urlopen(url) as url:
		data = json.loads(url.read())

	return data

def download_allen_ontologies():
	mouse = get_json_from_api(MOUSE_API)
	human = get_json_from_api(BRAINSPAN_API)

	#uncomment to dump json to file
	# with open('source/mouse.json', 'w') as outfile:
	# 	json.dump(mouse, outfile, indent=2)

	# with open('source/human.json', 'w') as outfile:
	# 	json.dump(human, outfile, indent=2)

	return mouse, human

def main():
	print("Running uberon cross-species mapping...")
	sparql_query = SparqlQueries(OWL_PATH, OWL_URI)

	#load in the spreadsheet mapping data
	mapping_sheet = get_report_data()

	#pull in the allen ontologies
	mouse_ontology, human_ontology = download_allen_ontologies()

	allen_institute_structures = AllenInstituteStructures(mapping_sheet, mouse_ontology, human_ontology)

	#write the output json file
	allen_institute_structures.write_output_json(OUTPUT_PATH, sparql_query)

	#write the output graph
	allen_institute_structures.write_output_graph(OUTPUT_GRAPH_PATH, MOUSE_OUTPUT_GRAPH_PATH, HUMAN_OUTPUT_GRAPH_PATH)

	print('finished...')

if __name__ == "__main__":
	main()
